import openpyxl
from pathlib import Path

path = Path(r"c:\Users\Lenovo Loq\Downloads\Vat Bill\Purchase Sales - BS Int 82-83.xlsx")
wb = openpyxl.load_workbook(path, data_only=True)
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('\nSheet:', name)
    max_row = min(ws.max_row, 20)
    max_col = min(ws.max_column, 20)
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
        rows.append(row)
    for r in rows:
        print(r)
